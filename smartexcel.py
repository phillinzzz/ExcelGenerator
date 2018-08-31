from openpyxl import workbook, load_workbook
from openpyxl.styles import Alignment, PatternFill, colors, Fill, Font
from datetime import datetime, date, timedelta
from calendar import monthrange


RED_COLOR = 'ff0000'
YELLOW_COLOR = 'ffff00'
WHITE_COLOR = 'ffffff'

def driver_generator():
    while True:
        yield ("2/3", WHITE_COLOR)
        yield ('0', WHITE_COLOR)
        yield ('2', WHITE_COLOR)


def zhizhang_generator():
    while True:
        yield ("2/3", RED_COLOR)
        yield ('B', WHITE_COLOR)
        yield ('0', YELLOW_COLOR)
        yield ('0', YELLOW_COLOR)


def zhibanyuan_generator_3_person():
    while True:
        yield ("2/3", RED_COLOR)
        yield ('B1', WHITE_COLOR)
        yield ('0', YELLOW_COLOR)
        yield ('0', YELLOW_COLOR)
        yield ("2/3", RED_COLOR)
        yield ('B2', WHITE_COLOR)
        yield ('0', YELLOW_COLOR)
        yield ('0', YELLOW_COLOR)
        yield ("2/3", RED_COLOR)
        yield ('B3', WHITE_COLOR)
        yield ('0', YELLOW_COLOR)
        yield ('0', YELLOW_COLOR)


def zhibanyuan_generator_4_person():
    while True:
        yield ("2/3", RED_COLOR)
        yield ('B1', WHITE_COLOR)
        yield ('0', YELLOW_COLOR)
        yield ('0', YELLOW_COLOR)
        yield ("2/3", RED_COLOR)
        yield ('B2', WHITE_COLOR)
        yield ('0', YELLOW_COLOR)
        yield ('0', YELLOW_COLOR)
        yield ("2/3", RED_COLOR)
        yield ('B3', WHITE_COLOR)
        yield ('0', YELLOW_COLOR)
        yield ('0', YELLOW_COLOR)
        yield ("2/3", RED_COLOR)
        yield ('B4', WHITE_COLOR)
        yield ('0', YELLOW_COLOR)
        yield ('0', YELLOW_COLOR)


class SmartExcel:

    oldFileName = "变电运维组轮值表2018年.xlsx"
    newFileName = "新变电运维组轮值表2018年.xlsx"
    # 各个值的人数（除正直长以外）
    groupInfo = {'group_A': (4, 3), 'group_B': (3, 3), 'group_C': (3, 4), 'group_D': (3, 3)}


    def __init__(self):

        # 司机的数量，需要手动设置
        self.driver_number = 6

        #复制表格
        self.old_wb = load_workbook(self.oldFileName)
        self.old_wb.save(self.newFileName)

        self.new_wb = load_workbook(self.newFileName)
        self.old_ws1 = self.new_wb["18-081"]
        self.old_ws2 = self.new_wb["18-082"]

        self.new_ws1 = self.new_wb.copy_worksheet(self.old_ws1)
        self.new_ws1.title = "18-091"

        self.new_ws2 = self.new_wb.copy_worksheet(self.old_ws2)
        self.new_ws2.title = "18-092"

        #除司机以外的白班人数
        self.employee_number = self.new_ws1.max_row - self.driver_number - 6

        # 上个月的月份
        self.oldStartDate = self.new_ws1['A2'].value
        self.oldEndDate = self.oldStartDate.replace(day=monthrange(self.oldStartDate.year, self.oldStartDate.month)[1])
        self.oldDate = (self.oldStartDate, self.oldEndDate)
        # print(self.oldDate)

        # 上月总天数
        daysOfOldMonth = monthrange(self.oldStartDate.year, self.oldStartDate.month)[1]

        # 获取司机上月最后一天的值班情况[2/3, 0, 2, 2/3, 0, 2]
        self.driver_status_last_day = []
        for driverID in range(1, self.driver_number+1):
            self.driver_status_last_day.append(str(self.new_ws1.cell(self.employee_number+5+driverID, daysOfOldMonth+3).value))

        # 值班员总数（包括值长）
        self.zhibanyuan_number = self.new_ws2.max_row - 6

        # 获取值班员（包括值长）上月最后1轮值的值班情况
        self.zhibanyuan_status_last_turn = []
        for zhibanyuanID in range(1, self.zhibanyuan_number+1):
            one_zhibanyuan_status_last_turn = []
            for i in range(-3, 1):
                one_zhibanyuan_status_last_turn.append(str(self.new_ws2.cell(zhibanyuanID+5, daysOfOldMonth+3+i).value))
            self.zhibanyuan_status_last_turn.append(one_zhibanyuan_status_last_turn)
        # print(self.zhibanyuan_status_last_turn)

        # 生成下个月的第一天以及最后一天。self.newDate是一个tuple.
        self.newDate = self.generateNextMonth(self.oldStartDate)

    def save(self):
        self.new_wb.save(self.newFileName)

    # 清空时间及数据
    def clear(self):
        # clear first sheet
        for i in range(3, self.new_ws1.max_row):
            for j in range(4, self.new_ws1.max_column+1):
                self.new_ws1.cell(i,j).value = ""
                self.new_ws1.cell(i,j).fill = PatternFill(start_color='ffffff', end_color='ffffff', fill_type="solid")
        # clear second sheet
            for i in range(3, self.new_ws2.max_row):
                for j in range(4, self.new_ws2.max_column + 1):
                    self.new_ws2.cell(i, j).value = ""
                    self.new_ws2.cell(i, j).fill = PatternFill(start_color='ffffff', end_color='ffffff', fill_type="solid")

    #修改一个表格的日期,此函数是给下一个函数用的
    def __generateFrame__(self, ws):

        # change title month
        ws['A2'].value = self.newDate[0]

        #update dates and weekdays
        start_date, end_date = self.newDate

        day_row_number = 3
        week_row_number = 4
        column_start_number = 4

        tmp_date = start_date
        column_number = column_start_number
        while tmp_date <= end_date:
            #填写日期行
            ws.cell(day_row_number, column_number).value = tmp_date
            ws.cell(day_row_number, column_number).number_format = 'DD'
            #填写周几行
            ws.cell(week_row_number, column_number).value = tmp_date
            ws.cell(week_row_number, column_number).number_format = '[$-804]aaa;@'

            tmp_date = tmp_date + timedelta(days=1)
            column_number += 1

    #将两个表格的日期修改好
    def generateFrames(self):
        self.__generateFrame__(self.new_ws1)
        self.__generateFrame__(self.new_ws2)

    # 计算下个月的月份（newDate）以及从1号到改月最后一天的range
    # 返回值：(newStartDate, newEndDate)
    def generateNextMonth(self, oldDate):
        oldMonth = oldDate.month
        oldYear = oldDate.year

        if oldMonth == 12:
            newYear = oldYear + 1
            newMonth = 1
        else:
            newYear = oldYear
            newMonth = oldMonth + 1

        newStartDate = date(year=newYear, month=newMonth, day=1)


        # newDate_year = newStartDate.year
        # newDate_month = newStartDate.month
        newEndDate_day = monthrange(newYear, newMonth)[1]
        newEndDate = date(year=newYear, month=newMonth, day=newEndDate_day)
        return (newStartDate, newEndDate)


    def tackle_first_sheet(self):
        # settle down the employees
        for employeeID in range(1, self.employee_number+1):

            # print(self.new_ws1.cell(employeeID+5, 3).value)

            column_start_number = 4
            column_number = column_start_number

            startDate = self.newDate[0]
            endDate = self.newDate[1]

            tmp_date = startDate


            while tmp_date <= endDate:
                #如果是周末
                if tmp_date.weekday() > 4:
                    self.new_ws1.cell(employeeID+5, column_number).value = '0'
                    self.new_ws1.cell(employeeID+5, column_number).alignment = Alignment(horizontal='center')
                    # self.new_ws1.cell(employeeID + 5, column_number).font = Font(color=colors.RED)
                    self.new_ws1.cell(employeeID+5, column_number).fill = PatternFill(start_color='ffff00',
                                                                                        end_color='ffff00',
                                                                                        fill_type="solid")
                else:
                    self.new_ws1.cell(employeeID+5, column_number).value = '2'
                    self.new_ws1.cell(employeeID+5, column_number).alignment = Alignment(horizontal='center')
                    # ws.cell(day_row_number, column_number).number_format = 'DD'
                    self.new_ws1.cell(employeeID+5, column_number).fill = PatternFill(start_color= 'ffffff',
                                                                                        end_color= 'ffffff',
                                                                                        fill_type="solid")

                tmp_date = tmp_date + timedelta(days=1)
                column_number += 1


        # now its time to arrange for the drivers
        for driverID in range(1, self.driver_number+1):

            column_start_number = 4
            column_number = column_start_number

            startDate = self.newDate[0]
            endDate = self.newDate[1]

            tmp_date = startDate
            g = driver_generator()
            i = 0
            # 首先调整发生器知道其输出与上个月最后一天一致
            # print(self.driver_status_last_day)
            while (next(g)[0] != self.driver_status_last_day[driverID-1]) & (i < 4):
                i += 1
                # print(i)

            while tmp_date <= endDate:
                cell_data = next(g)
                self.new_ws1.cell(driverID + self.employee_number+5, column_number).value = cell_data[0]
                self.new_ws1.cell(driverID + self.employee_number+5, column_number).alignment = Alignment(horizontal='center')
                self.new_ws1.cell(driverID + self.employee_number+5, column_number).fill = PatternFill(start_color=cell_data[1],
                                                                                    end_color=cell_data[1],
                                                                                    fill_type="solid")
                # self.new_ws1.cell(driverID + self.employee_number+5, column_number).fill = PatternFill(start_color='ffffff',
                #                                                                    end_color='ffffff',
                #                                                                    fill_type="solid")
                # self.new_ws1.cell(employeeID + 5, column_number).font = Font(color=colors.RED)

                # if self.new_ws1.cell(driverID + self.employee_number+5, column_number).value == "\'2\/3":
                #     self.new_ws1.cell(driverID + self.employee_number+5, column_number).number_format = '# ?/?'


                tmp_date = tmp_date + timedelta(days=1)
                column_number += 1


    def tackle_second_sheet(self):

        for zhibanyuanID  in range(1, self.zhibanyuan_number+1):
            print("zhibanyuan: %s enter first step!" %(zhibanyuanID))
            print(self.zhibanyuan_status_last_turn[zhibanyuanID-1])
            column_start_number = 4
            column_number = column_start_number

            startDate = self.newDate[0]
            endDate = self.newDate[1]

            tmp_date = startDate

            # 设置以及调整发生器
            g = self.generatorSetter(zhibanyuanID)

            while tmp_date <= endDate:
                cell_data = next(g)
                self.new_ws2.cell(zhibanyuanID+5, column_number).value = cell_data[0]
                self.new_ws2.cell(zhibanyuanID+5, column_number).alignment = Alignment(horizontal='center')
                self.new_ws2.cell(zhibanyuanID+5, column_number).fill = PatternFill(start_color=cell_data[1],
                                                                                    end_color=cell_data[1],
                                                                                    fill_type="solid")
                # self.new_ws1.cell(employeeID + 5, column_number).font = Font(color=colors.RED)

                # if self.new_ws1.cell(driverID + self.employee_number+5, column_number).value == "\'2\/3":
                #     self.new_ws1.cell(driverID + self.employee_number+5, column_number).number_format = '# ?/?'


                tmp_date = tmp_date + timedelta(days=1)
                column_number += 1




    def generatorSetter(self, zhibanyuanID):

        #确定值班员类型identity（值长/四人值值班员/三人值值班员）
        identity = ''
        #发生器容器，根据值班员类型确定
        g = ''

        column_start_number = 4
        column_number = column_start_number
        startDate = self.oldDate[0]
        endDate = self.oldDate[1]
        tmp_date = startDate

        while tmp_date <= endDate:
            # print("entering loop!")
            #
            # print(self.old_ws2.cell(zhibanyuanID+5, column_number).value)
            if self.old_ws2.cell(zhibanyuanID+5, column_number).value == 'B':
                identity = 'zhizhang'
                g = zhizhang_generator()
                break
            elif self.old_ws2.cell(zhibanyuanID+5, column_number).value == 'B4':
                identity = 'zhibanyuan_4'
                g = zhibanyuan_generator_4_person()
                break
            tmp_date = tmp_date + timedelta(days=1)
            column_number += 1

        if identity == '':
            identity = 'zhibanyuan_3'
            g = zhibanyuan_generator_3_person()

        print("the %s person is a " %self.new_ws2.cell(zhibanyuanID+5, 3).value + identity)
        print("and his last turn is: ")
        print(self.zhibanyuan_status_last_turn[zhibanyuanID-1])

        #调整发生器的输出，与该值班员最后一个BX对应上即可。
        if identity == 'zhizhang':
            for i in range(1, 5):
                if self.zhibanyuan_status_last_turn[zhibanyuanID - 1][i - 1] == 'B':
                    for j in range(0, 6-i):
                        next(g)
                    break
        elif identity == 'zhibanyuan_4':
            for i in range(1, 5):
                if self.zhibanyuan_status_last_turn[zhibanyuanID - 1][i - 1] == 'B1':
                    for j in range(0, 6-i):
                        next(g)
                    break
                elif self.zhibanyuan_status_last_turn[zhibanyuanID - 1][i - 1] == 'B2':
                    for j in range(0, 10-i):
                        next(g)
                    break
                elif self.zhibanyuan_status_last_turn[zhibanyuanID - 1][i - 1] == 'B3':
                    for j in range(0, 14-i):
                        next(g)
                    break
                elif self.zhibanyuan_status_last_turn[zhibanyuanID - 1][i - 1] == 'B4':
                    for j in range(0, 18-i):
                        next(g)
                    break
        else:
            for i in range(1, 5):
                if self.zhibanyuan_status_last_turn[zhibanyuanID - 1][i - 1] == 'B1':
                    for j in range(0, 6-i):
                        next(g)
                    break
                elif self.zhibanyuan_status_last_turn[zhibanyuanID - 1][i - 1] == 'B2':
                    for j in range(0, 10-i):
                        next(g)
                    break
                elif self.zhibanyuan_status_last_turn[zhibanyuanID - 1][i - 1] == 'B3':
                    for j in range(0, 14-i):
                        next(g)
                    break


        return g








if __name__ == "__main__":
    smartExcel = SmartExcel()
    smartExcel.clear()
    smartExcel.generateFrames()


    smartExcel.tackle_first_sheet()
    # print(smartExcel.zhibanyuan_status_last_turn[23])
    # print(smartExcel.zhibanyuan_status_last_turn[23][3])
    # g = zhizhang_generator()
    # for i in range(1,5):
    #     print(next(g)==smartExcel.zhibanyuan_status_last_turn[23][3])

    smartExcel.tackle_second_sheet()
    # smartExcel.new_ws1['AI7'].fill = PatternFill(start_color='ffff00',end_color='ffff00',fill_type="solid")


    # g = smartExcel.generatorSetter(1)
    # for i in range(1, 5):
    #     print(next(g))

    # print(smartExcel.old_ws2.cell(6, 6).value)
    smartExcel.save()